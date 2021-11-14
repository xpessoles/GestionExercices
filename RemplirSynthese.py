# -*- coding: utf-8 -*-
"""
Created on Fri Nov 12 18:39:35 2021

@author: xpessoles
Remplir le fichier de Synthèse des exercices (commençons par les DDS)
"""

"""
1. Donner la liste des dossiers contenant des exos
2. Dans ces dossiers, rechercher tous les exos
3. Pour un exercice, remplir une ligne du tableau
"""

import os
import openpyxl

dossiers_exos = ["C:\GitHub\ExercicesCompetences"]
fichier_synthese = "SyntheseExercices.xlsx"
fichier_competences = "CompetencesCPGE2021.xlsx"
filiere = 'PCSI-PSI'
path = "C:\GitHub\ExercicesCompetences"

exo = 'C:\\GitHub\\ExercicesCompetences\\A3_AnalyseFonctionnelleStructurelle\\A3_01_ChaineFonctionnelle\\58_Oz440\\58_Oz440.tex'

def rechercher_exos(dossiers_exos:list) -> list : 
    """
    Parameters
    ----------
    dossiers_exos : list
        Liste des chemins absolus des dossiers comprenant des exercices.

    Returns
    -------
    list 
        Liste des chemins absolus de tous les fichiers tex.

    """
    file_list = []
    for path in dossiers_exos : 
        liste_exos = rechercher_exos_path(path)
        for exo in liste_exos : 
            file_list.append(exo)
    return file_list
    
def rechercher_exos_path(path:str) -> list : 
    """
    Parameters
    ----------
    path : str
        Chemin absolu d'un comprenant des exercices.

    Returns
    -------
    list 
        Liste des chemins absolus de tous les fichiers tex contenu dans le path.
    """
    file_list =[]
    for path, folders, files in os.walk(path):
        for file in files:
            if '.tex' in file : 
                file_list.append(os.path.join(path, file))
    return file_list
    
#lf = rechercher_exos(dossiers_exos)

def rechercher_balises(exo:str) -> list : 
    """
    Trouver les balises dans un fichier tex
    Parameters
    ----------
    exo : str
        Chemin vers un exercice.

    Returns
    -------
    list 
        [path,nom,[type],corrige,[classes],semestre,[competences]]
    """
    path,nom,type_exo,corrige,classes,semestre,competences = exo,None,[],None,[],None,[]
    
    nom = exo.split('\\')[-1][:-4]
    
    fid = open(exo,'r')
    data = fid.readlines()
    fid.close()
    for ligne in data : 
        # Types d'exos
        if "ddstrue" in ligne : 
            type_exo.append("dds")
        if "tdtrue" in ligne : 
            type_exo.append("td")
        if "colletrue" in ligne : 
            type_exo.append("colle")
        if "applicationtrue" in ligne : 
            type_exo.append("application")
        if "activationtrue" in ligne : 
            type_exo.append("activation")
        
        # Correction
        if "correctionfalse" in ligne : 
            corrige = True
        if "correctionfalse" in ligne : 
            corrige = False
        
        # Classes
        # Voir avec macro UPSTI
        
        # Semestre
        # A lier à la compétence
        
        # Competences
        if "UPSTIcompetence" in ligne :
            l = ligne.split("{")
            l = l[1]
            l = l.split("}")
            competences.append(l[0])
            
    res = [path,nom,type_exo,corrige,classes,semestre,competences]
    return res


def lire_fichier_competences(fichier,filiere):
    """
    Création d'un dictionnaire des compétences d'une filière
    """
    wb = openpyxl.load_workbook(fichier, data_only = True)# charge le workbook en donnant les valeurs plutôt que les formules (le défaut est data_only = False).
    
    comp = {}
    filiere = "PCSI-PSI"
    onglet = filiere
    feuille = wb[onglet]
    for i in range(1,feuille.max_row+1):
        cell='A'+str(i)
        val = feuille[cell].value
        if (type(val)!= type(None)) and ("-" in val) :
            comp[val]=None
    wb.close()
    return comp

def lire_fichier_synthese(fichier,competences):
    """
    Association de la colonne à la compétence
    """
    wb = openpyxl.load_workbook(fichier, data_only = True)# charge le workbook en donnant les valeurs plutôt que les formules (le défaut est data_only = False).
    

    onglet = "BilanExos"
    feuille = wb[onglet] # Sélection de la feuille
    
    for col in feuille.iter_cols():
        for cell in col :
            print(cell.value)
    wb.close()


comp = lire_fichier_competences(fichier_competences,filiere)
lire_fichier_synthese(fichier_synthese,comp)
